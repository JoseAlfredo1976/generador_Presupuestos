"""
Generates the valoración Excel: one sheet with all line items, totals and IVA.
"""
from pathlib import Path

import openpyxl
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side,
                              numbers)
from openpyxl.utils import get_column_letter

BLUE = "4F84B5"
WHITE = "FFFFFF"
LIGHT_BLUE = "D6E4F0"
DARK = "1F1F1F"


class ExcelGenerator:
    def generate(self, output_path: Path, data: dict, items: list[dict]):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Valoración"

        self._set_column_widths(ws)
        self._write_header(ws, data)
        last_row = self._write_items(ws, items)
        self._write_totals(ws, last_row, items)

        wb.save(output_path)

    # ------------------------------------------------------------------
    def _set_column_widths(self, ws):
        ws.column_dimensions["A"].width = 60
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 16

    def _header_font(self, bold=False):
        return Font(name="Calibri", size=11, bold=bold, color=WHITE)

    def _body_font(self, bold=False, color=DARK):
        return Font(name="Calibri", size=10, bold=bold, color=color)

    def _thin_border(self):
        s = Side(style="thin", color="AAAAAA")
        return Border(left=s, right=s, top=s, bottom=s)

    # ------------------------------------------------------------------
    def _write_header(self, ws, data: dict):
        # Title row
        ws.merge_cells("A1:E1")
        title = ws["A1"]
        title.value = f"VALORACIÓN DE OBRA – Presupuesto {data.get('[[CONTRATO_NUM]]', '')}"
        title.font = Font(name="Calibri", size=14, bold=True, color=BLUE)
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # Info rows
        info = [
            ("Obra:", data.get("[[OBRA_COMUNIDAD]]", "")),
            ("Cliente:", data.get("[[CLIENTE_NOMBRE]]", "")),
            ("Fecha:", data.get("[[FECHA_CONTRATO]]", "")),
        ]
        for i, (label, value) in enumerate(info, start=2):
            ws.cell(row=i, column=1, value=label).font = Font(name="Calibri", size=10, bold=True)
            ws.cell(row=i, column=2, value=value).font = Font(name="Calibri", size=10)
            ws.merge_cells(f"B{i}:E{i}")
        ws.row_dimensions[5].height = 8  # spacer

        # Column headers
        headers = ["DESCRIPCIÓN", "UD", "UDS", "PRECIO (€)", "IMPORTE (€)"]
        fill = PatternFill("solid", fgColor=BLUE)
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=6, column=col, value=h)
            cell.font = self._header_font(bold=True)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self._thin_border()
        ws.row_dimensions[6].height = 20

    def _write_items(self, ws, items: list[dict]) -> int:
        fills = [None, PatternFill("solid", fgColor=LIGHT_BLUE)]
        for i, item in enumerate(items):
            row = 7 + i
            fill = fills[i % 2]

            desc_cell = ws.cell(row=row, column=1, value=item["descripcion"])
            desc_cell.font = self._body_font()
            desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                desc_cell.fill = fill

            ud_cell = ws.cell(row=row, column=2, value=item["unidad"])
            ud_cell.font = self._body_font()
            ud_cell.alignment = Alignment(horizontal="center")
            if fill:
                ud_cell.fill = fill

            qty_cell = ws.cell(row=row, column=3, value=item["cantidad"])
            qty_cell.font = self._body_font()
            qty_cell.alignment = Alignment(horizontal="right")
            qty_cell.number_format = "#,##0.00"
            if fill:
                qty_cell.fill = fill

            price_cell = ws.cell(row=row, column=4, value=item["precio_unitario"])
            price_cell.font = self._body_font()
            price_cell.alignment = Alignment(horizontal="right")
            price_cell.number_format = "#,##0.00 €"
            if fill:
                price_cell.fill = fill

            total_cell = ws.cell(row=row, column=5, value=item["importe"])
            total_cell.font = self._body_font()
            total_cell.alignment = Alignment(horizontal="right")
            total_cell.number_format = "#,##0.00 €"
            if fill:
                total_cell.fill = fill

            for col in range(1, 6):
                ws.cell(row=row, column=col).border = self._thin_border()

            ws.row_dimensions[row].height = max(30, len(item["descripcion"]) // 6)

        return 6 + len(items)

    def _write_totals(self, ws, last_row: int, items: list[dict]):
        total_sin_iva = sum(i["importe"] for i in items)
        iva_amount = total_sin_iva * 0.21
        total_con_iva = total_sin_iva + iva_amount

        rows = [
            ("TOTAL PRESUPUESTO (sin IVA)", total_sin_iva),
            ("IVA (21%)", iva_amount),
            ("TOTAL CON IVA", total_con_iva),
        ]
        fill_total = PatternFill("solid", fgColor=BLUE)
        for offset, (label, amount) in enumerate(rows):
            row = last_row + 1 + offset
            ws.merge_cells(f"A{row}:D{row}")
            label_cell = ws.cell(row=row, column=1, value=label)
            amount_cell = ws.cell(row=row, column=5, value=amount)
            if label == "TOTAL CON IVA":
                label_cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
                amount_cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
                label_cell.fill = fill_total
                amount_cell.fill = fill_total
            else:
                label_cell.font = self._body_font(bold=True)
                amount_cell.font = self._body_font(bold=True)
            label_cell.alignment = Alignment(horizontal="right", vertical="center")
            amount_cell.alignment = Alignment(horizontal="right", vertical="center")
            amount_cell.number_format = "#,##0.00 €"
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = self._thin_border()
