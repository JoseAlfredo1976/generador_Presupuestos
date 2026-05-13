"""
Base de datos local: clientes y contadores de numeracion.
Fuente: tarifas/BASE_DATOS.xlsx
"""
from __future__ import annotations

import logging
from pathlib import Path

log = logging.getLogger(__name__)

BASE_PATH = Path(__file__).parent.parent / "tarifas" / "BASE_DATOS.xlsx"

# Tipos que comparten numero con el presupuesto asociado (no tienen contador propio)
# El usuario introduce el numero manualmente (igual que el presupuesto de la misma obra)
TIPOS_SIN_CONTADOR = {
    "cctv_bajante",       # mismo numero que el presupuesto de limpieza/obra
    "inspeccion_zum",     # mismo numero que el presupuesto de limpieza
    "informe_desatasco",  # mismo numero que el presupuesto de desatasco
    "bajantes_amianto",   # mismo numero que el presupuesto de obra
    "certificado_obra",   # mismo numero que el presupuesto de obra
    "contrato_subcontrata",  # mismo numero que el presupuesto de la misma calle
    "plan_seguridad",     # mismo numero que el presupuesto de obra asociado
}

# Mapeo tipo -> CATEGORIA en hoja CONTROL del Excel
TIPO_A_CATEGORIA: dict[str, str] = {
    "obra_1":               "POCERIA",
    "obra_2":               "POCERIA",
    "desatasco":            "POCERIA",
    "fuga_agua":            "POCERIA",
    "limpieza_aerea":       "LIMPIEZAS",
    "fresador":             "LIMPIEZAS",
    "robot_limpieza":       "LIMPIEZAS",
    "vaciado_fosa":         "LIMPIEZAS",
    "contrato_saneamiento": "CONTRATOS DE MANTENIMIENTO",
    "fontaneria":           "FONTANERIA",
    "albanileria":          "ALBANILERIA Y OTROS TRABAJOS",
}

# Overrides manuales cuando el BASE_DATOS no refleja el numero correcto
OVERRIDES_SIGUIENTE: dict[str, int] = {
    "CONTRATOS DE MANTENIMIENTO": 2221,
}

_clientes: list[str] | None = None
_contadores: dict[str, int] | None = None


def _load():
    global _clientes, _contadores
    if _clientes is not None:
        return
    if not BASE_PATH.exists():
        log.warning("BASE_DATOS.xlsx no encontrado en %s", BASE_PATH)
        _clientes = []
        _contadores = {}
        return

    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(BASE_PATH), read_only=True, data_only=True)

        # Clientes
        ws_cli = wb["CLIENTES"]
        nombres = []
        for i, row in enumerate(ws_cli.iter_rows(values_only=True)):
            if i == 0:
                continue  # cabecera
            nombre = row[0]
            if nombre and str(nombre).strip():
                nombres.append(str(nombre).strip())
        _clientes = sorted(set(nombres))

        # Contadores desde hoja CONTROL
        ws_ctrl = wb["CONTROL"]
        contadores = {}
        for i, row in enumerate(ws_ctrl.iter_rows(values_only=True)):
            if i == 0:
                continue
            categoria, _, _, _, _, siguiente = row[:6]
            if categoria and siguiente:
                try:
                    contadores[str(categoria).strip()] = int(siguiente)
                except (ValueError, TypeError):
                    pass
        _contadores = contadores

        log.info("BASE_DATOS cargada: %d clientes, %d categorias", len(_clientes), len(_contadores))
    except Exception as e:
        log.warning("Error cargando BASE_DATOS: %s", e)
        _clientes = []
        _contadores = {}


def buscar_clientes(q: str, limit: int = 20) -> list[str]:
    """Devuelve clientes cuyo nombre contiene q (case-insensitive)."""
    _load()
    if not q or not _clientes:
        return []
    q_lower = q.lower()
    return [c for c in _clientes if q_lower in c.lower()][:limit]


def siguiente_numero_local(tipo: str) -> int | None:
    """
    Devuelve el siguiente numero para el tipo dado segun la base local.
    Retorna None si el tipo no tiene contador propio (TIPOS_SIN_CONTADOR)
    o si no hay datos disponibles.
    """
    if tipo in TIPOS_SIN_CONTADOR:
        return None
    _load()
    if not _contadores:
        return None
    cat = TIPO_A_CATEGORIA.get(tipo)
    if not cat:
        return None
    # Override manual tiene prioridad sobre el valor del Excel
    if cat in OVERRIDES_SIGUIENTE:
        return OVERRIDES_SIGUIENTE[cat]
    return _contadores.get(cat)
